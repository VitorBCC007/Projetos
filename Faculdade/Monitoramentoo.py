import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.chart import Reference
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import locale
from openpyxl.chart import BarChart

locale.setlocale(locale.LC_ALL, 'pt_BR.utf-8')

class Principal(webdriver.Chrome):
    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])
        service = Service(ChromeDriverManager().install())
        super(Principal, self).__init__(service=service, options=options)
        self.maximize_window()

    def abrir_link(self, link):
        self.get(link)
        

workbook = openpyxl.load_workbook('C:/Users/Joao Vitor/Desktop/dadosapresentacao.xlsx')
worksheet = workbook[workbook.sheetnames[0]]

driver = Principal()

precos = []
nomes_produtos = []  

for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, values_only=True, min_col=1, max_col=2):
    link = row[0]
    nome_produto = row[1] if len(row) > 1 else ""

    if link is not None and isinstance(link, str):  # Verifica se o link é uma string válida
        driver.abrir_link(link)

        # PREÇOS
        if 'mercadolivre.com' in link:
            preco_element = driver.find_element(By.CLASS_NAME, 'andes-money-amount__fraction')
            preco = preco_element.text
            nome_produto_element = driver.find_element(By.CLASS_NAME, 'ui-pdp-title')
            nome_produto = nome_produto_element.text

        elif 'kabum.com' in link:
            try:
                preco_element = driver.find_element(By.XPATH, '//*[@id="blocoValores"]/div[2]/div[1]/h4')
                nome_produto_element = driver.find_element(By.XPATH, '//*[@id="__next"]/main/article/section/div[3]/div[1]/div/h1')
            except:
                nome_produto_element = driver.find_element(By.XPATH, '//*[@id="__next"]/main/article/section/div[2]/div[1]/div/h1')
            preco = preco_element.text
            nome_produto = nome_produto_element.text

        elif 'aliexpress.com' in link:
            try:                                               
                preco_element = driver.find_element(By.CLASS_NAME, 'uniform-banner-box-price')
                nome_produto_element = driver.find_element(By.CLASS_NAME, 'product-title-text')
            except:
                preco_element = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]/div[3]/div[1]')
                nome_produto_element = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/div/div[2]/div[1]/h1')

            preco = preco_element.text
            nome_produto = nome_produto_element.text

        elif 'amazon.com' in link:
            preco_element = driver.find_element(By.CLASS_NAME, 'a-price-whole')
            preco = preco_element.text
            nome_produto_element = driver.find_element(By.ID, 'productTitle')
            nome_produto = nome_produto_element.text


        else:
            preco = 'Preço não encontrado'

        if preco != 'Preço não encontrado':
            # remover espaços
            preco = preco.replace('R$', '').strip().replace('.', '')
            # , por .
            preco = preco.replace(',', '.')
            preco = float(preco)
        else:
            preco = 0.0  

        # Formatar o preço em Reais
        preco_formatado = locale.format_string('%.2f', preco, grouping=True)

        precos.append(preco)  # Salvando preço
        nomes_produtos.append(nome_produto)  # Salvando nome produto
    else:
        preco = 'Link inválido'
        precos.append(preco)
        nomes_produtos.append(nome_produto)

driver.quit()

# COLOCANDO PREÇO E NOME DO PRODUTO NA PLANILHA
nova_planilha = workbook.create_sheet("Preços dos Produtos")

cabecalhos = ['Link', 'Nome do Produto', 'Preço']
for col_num, cabecalho in enumerate(cabecalhos, start=1):
    col_letra = get_column_letter(col_num)
    nova_planilha[f'{col_letra}1'] = cabecalho
    nova_planilha[f'{col_letra}1'].alignment = Alignment(horizontal='center')
    nova_planilha[f'{col_letra}1'].font = Font(bold=True)

for row_num, (link, nome_produto, preco) in enumerate(zip(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, values_only=True, min_col=1, max_col=2), nomes_produtos, precos), start=2):
    nova_planilha[f'A{row_num}'] = link[0]
    nova_planilha[f'B{row_num}'] = nome_produto
    nova_planilha[f'C{row_num}'] = preco

# Aplicar formatação de célula e alinhamento
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

font = Font(name='Arial', size=12, bold=True)

for row in nova_planilha.iter_rows(min_row=1, max_row=len(precos) + 1):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        cell.font = font

nova_planilha.column_dimensions['A'].width = 23
nova_planilha.column_dimensions['B'].width = 190
nova_planilha.column_dimensions['C'].width = 15

# GRÁFICO
chart = BarChart()  
data = Reference(nova_planilha, min_col=3, min_row=1, max_row=len(precos) + 1, max_col=3)
categories = Reference(nova_planilha, min_col=2, min_row=2, max_row=len(precos) + 1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Variação de Preços"
chart.x_axis.title = "Produto"
chart.y_axis.title = "Preço"

chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 3000
chart.y_axis.majorUnit = 300

chart.width = 36
chart.height = 18

nova_planilha.add_chart(chart, "E3")

workbook.save("C:/Users/Joao Vitor/Desktop/preco_ordem.xlsx")
